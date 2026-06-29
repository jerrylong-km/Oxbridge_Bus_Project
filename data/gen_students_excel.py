"""
把 virtual_students.json 的 100 条记录转为 Excel 文件。
列：序号 / 姓名(student_id) / 年级(6-12随机) / 地址(Google Geocoding 反查)
运行：python gen_students_excel.py
输出：virtual_students.xlsx（同目录）
"""
import json
import random
import time
import os
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 配置 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
JSON_PATH   = os.path.join(SCRIPT_DIR, "virtual_students.json")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "virtual_students.xlsx")
API_KEY     = "YOUR_GOOGLE_MAPS_API_KEY_HERE"   # 填入你的 backend key（无 referer 限制）
DELAY_SEC   = 0.12   # 每次 Geocoding 请求间隔，避免触发速率限制
RANDOM_SEED = 42
# ─────────────────────────────────────────────────────────────────────────────

random.seed(RANDOM_SEED)

def reverse_geocode(lat, lng, api_key):
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"latlng": f"{lat},{lng}", "key": api_key}
    try:
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()
        if data.get("status") == "OK" and data.get("results"):
            return data["results"][0]["formatted_address"]
    except Exception as e:
        print(f"  [警告] 请求异常: {e}")
    return f"{lat},{lng}"   # 失败时回退为坐标字符串

def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        students = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # 表头样式
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["序号", "姓名", "年级", "地址"]
    col_widths = [8, 14, 8, 60]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.row_dimensions[1].height = 20

    total = len(students)
    for i, stu in enumerate(students, start=1):
        grade = random.randint(6, 12)
        print(f"[{i:3}/{total}] {stu['student_id']}  反查地址中…", end=" ", flush=True)
        address = reverse_geocode(stu["lat"], stu["lng"], API_KEY)
        print(address[:60])

        ws.append([i, stu["student_id"], grade, address])
        # 数据行居中（序号、姓名、年级）
        for col in range(1, 4):
            ws.cell(row=i + 1, column=col).alignment = Alignment(horizontal="center")

        time.sleep(DELAY_SEC)

    wb.save(OUTPUT_PATH)
    print(f"\n✅ 已保存：{OUTPUT_PATH}")

if __name__ == "__main__":
    main()
