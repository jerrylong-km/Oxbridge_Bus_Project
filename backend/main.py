# main.py
import os
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from sqlalchemy.orm import Session
from passlib.context import CryptContext
import models
import schemas
from database import engine, get_db
from auth_deps import get_current_user, require_role
from auth_utils import get_password_hash, verify_password

# --- 路径基准：以本文件所在的 backend/ 目录为锚点，与启动时的工作目录无关 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))          # .../Oxbridge_Bus_Project/backend
PROJECT_ROOT = os.path.dirname(BASE_DIR)                        # .../Oxbridge_Bus_Project
FRONTEND_DIR = os.path.join(PROJECT_ROOT, "frontend")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# 建表（如果尚不存在）
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="智慧校车多租户 SaaS 平台 API")

# --- 注册认证路由 ---
from auth_router import router as auth_router
app.include_router(auth_router)

# 处理 Chrome Private Network Access 预检请求（必须在 CORSMiddleware 外层）
class PrivateNetworkAccessMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        # 拦截带有 PNA 头的 OPTIONS 预检，直接返回正确响应
        if (request.method == "OPTIONS" and
                request.headers.get("access-control-request-private-network")):
            from starlette.responses import Response
            origin = request.headers.get("origin", "")
            req_headers = request.headers.get("access-control-request-headers", "content-type")
            response = Response(status_code=200)
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            response.headers["Access-Control-Allow-Headers"] = req_headers
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Private-Network"] = "true"
            response.headers["Access-Control-Max-Age"] = "600"
            return response
        response = await call_next(request)
        # 对 HTML/JS 响应禁止浏览器缓存，确保鉴权逻辑始终是最新版本
        path = request.url.path
        if path.endswith('.html') or path.endswith('.js'):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
            response.headers["Pragma"] = "no-cache"
        return response

# CORSMiddleware 先加（内层），PrivateNetworkAccessMiddleware 后加（外层，先执行）
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5500",
        "http://localhost:5500"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(PrivateNetworkAccessMiddleware)

# 密码加密上下文
pwd_context = CryptContext(schemes=["sha256_crypt"], deprecated="auto")

# 基础测试 API
@app.get("/")
def read_root():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/login.html")

# --- Token 验证端点：前端用来确认 token 是否仍然有效 ---
@app.get("/api/me")
def get_me(current_user: models.User = Depends(get_current_user)):
    return {"user_id": current_user.user_id, "username": current_user.username, "role": current_user.role}

# HTML 页面路由（前端文件位于项目根目录的 frontend/ 内）
@app.get("/register.html", response_class=FileResponse)
def serve_register():
    return FileResponse(os.path.join(FRONTEND_DIR, "register.html"))

@app.get("/index.html", response_class=FileResponse)
def serve_index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

# --- 登录页面 ---
@app.get("/login.html")
def serve_login():
    return FileResponse(os.path.join(FRONTEND_DIR, "login.html"), headers={"Cache-Control": "no-store"})

# --- 超級管理員審核工作台頁面 ---
@app.get("/auth_guard.js")
def serve_auth_guard():
    return FileResponse(os.path.join(FRONTEND_DIR, "auth_guard.js"), headers={"Cache-Control": "no-store"})

@app.get("/i18n.js")
def serve_i18n():
    return FileResponse(os.path.join(FRONTEND_DIR, "i18n.js"), headers={"Cache-Control": "no-store"})

@app.get("/admin_dashboard.html")
def serve_admin_dashboard():
    return FileResponse(os.path.join(FRONTEND_DIR, "admin_dashboard.html"), headers={"Cache-Control": "no-store"})

# --- SchoolAdmin 校车线路管理页 ---
@app.get("/school_dashboard.html")
def serve_school_dashboard():
    return FileResponse(os.path.join(FRONTEND_DIR, "school_dashboard.html"), headers={"Cache-Control": "no-store"})

# --- 其他前端页面 ---
@app.get("/actual_routes.html", response_class=FileResponse)
def serve_actual_routes():
    return FileResponse(os.path.join(FRONTEND_DIR, "actual_routes.html"))

@app.get("/preview_stops.html", response_class=FileResponse)
def serve_preview_stops():
    return FileResponse(os.path.join(FRONTEND_DIR, "preview_stops.html"))

@app.get("/preview_students.html", response_class=FileResponse)
def serve_preview_students():
    return FileResponse(os.path.join(FRONTEND_DIR, "preview_students.html"))

@app.get("/preview_zones.html", response_class=FileResponse)
def serve_preview_zones():
    return FileResponse(os.path.join(FRONTEND_DIR, "preview_zones.html"))

# --- 【全新核心功能】：学校注册 API ---
@app.post("/api/register_school")
def register_school(request: schemas.SchoolRegister, db: Session = Depends(get_db)):
    # 1. 检查邮箱或用户名是否已经被注册
    existing_user = db.query(models.User).filter(
        (models.User.email == request.email) | (models.User.username == request.username)
    ).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="该邮箱或用户名已被注册")

    # 2. 将学校信息存入 Schools 表
    new_school = models.School(
        school_name=request.school_name,
        address=request.address,
        latitude=request.latitude,
        longitude=request.longitude,
        approval_status="待审核"  # 默认状态为待审核
    )
    db.add(new_school)
    db.commit()
    db.refresh(new_school) # 获取数据库自动生成的 school_id

    # 3. 对密码进行不可逆的 bcrypt 哈希加密（统一走 auth_utils，与登录校验保持一致）
    hashed_password = get_password_hash(request.password)

    # 4. 将管理员信息存入 Users 表，并绑定刚才生成的 school_id
    new_user = models.User(
        username=request.username,
        email=request.email,
        password_hash=hashed_password,
        role="SchoolAdmin",
        school_id=new_school.school_id
    )
    db.add(new_user)
    db.commit()

    return {"status": "success", "message": "注册申请已提交，请等待超级管理员审核！"}

# ---------------------------------------------------------
# 超級管理員審核工作台 API (SQLAlchemy ORM 版本)
# ---------------------------------------------------------
from pydantic import BaseModel
from typing import Optional

# 1. 定義前端傳來的審核決策數據結構
class ReviewDecision(BaseModel):
    school_id: int
    decision: str  # 支援 "已通过" 或 "已拒绝"
    reason: Optional[str] = None  # 拒絕理由（選填，批准時前端傳 null）

# 2. 獲取所有「待審核」學校清單的介面（需要 SuperAdmin 角色）
@app.get("/api/admin/pending_schools")
def get_pending_schools(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SuperAdmin")),
):
    # 完美複用 main.py 的 db 模式，查詢審核狀態為 "待审核" 的學校
    schools = db.query(models.School).filter(models.School.approval_status == "待审核").all()
    return schools

# 2b. 獲取「已通過 / 已暫停」的租戶清單（用於租戶管理面板）
@app.get("/api/admin/managed_schools")
def get_managed_schools(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SuperAdmin")),
):
    """回傳已開通（已通过）與已暫停（已暂停）的學校，並附上其管理員帳號資訊。"""
    schools = db.query(models.School).filter(
        models.School.approval_status.in_(["已通过", "已暂停"])
    ).all()

    result = []
    for s in schools:
        # 取該校的管理員帳號（SchoolAdmin），用於前端展示
        admin = db.query(models.User).filter(
            models.User.school_id == s.school_id,
            models.User.role == "SchoolAdmin",
        ).first()
        result.append({
            "school_id": s.school_id,
            "school_name": s.school_name,
            "address": s.address,
            "approval_status": s.approval_status,
            "admin_username": admin.username if admin else None,
            "admin_email": admin.email if admin else None,
        })
    return result

# 3. 提交審核結果（通過/駁回）的介面（需要 SuperAdmin 角色）
@app.post("/api/admin/review_school")
def review_school(
    payload: ReviewDecision,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SuperAdmin")),
):
    # 安全校驗：確保傳入的決策合法
    if payload.decision not in ["已通过", "已拒绝"]:
        raise HTTPException(status_code=400, detail="無效的審核決策，請傳入 '已通过' 或 '已拒绝'")

    # 尋找對應的學校記錄（相容 id 或 school_id 的欄位命名）
    school_field = models.School.id if hasattr(models.School, "id") else models.School.school_id
    school = db.query(models.School).filter(school_field == payload.school_id).first()
    
    if not school:
        raise HTTPException(status_code=404, detail="找不到該學校記錄")

    # 更新狀態與拒絕理由
    school.approval_status = payload.decision
    if payload.reason and hasattr(school, "rejection_reason"):
        school.rejection_reason = payload.reason

    db.commit()  # 提交事務
    return {"status": "success", "message": f"學校審核完成，結果已變更為：{payload.decision}"}

# 4. 暫停 / 恢復租戶（在「已通过」與「已暂停」之間切換）
class StatusToggle(BaseModel):
    school_id: int
    action: str  # "suspend"（暫停）或 "resume"（恢復）

@app.post("/api/admin/toggle_school_status")
def toggle_school_status(
    payload: StatusToggle,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SuperAdmin")),
):
    if payload.action not in ["suspend", "resume"]:
        raise HTTPException(status_code=400, detail="無效操作，請傳入 'suspend' 或 'resume'")

    school = db.query(models.School).filter(
        models.School.school_id == payload.school_id
    ).first()
    if not school:
        raise HTTPException(status_code=404, detail="找不到該學校記錄")

    if payload.action == "suspend":
        school.approval_status = "已暂停"
        msg = "已暫停該租戶，其管理員將無法登入"
    else:
        school.approval_status = "已通过"
        msg = "已恢復該租戶，其管理員可正常登入"

    db.commit()
    return {"status": "success", "message": msg, "new_status": school.approval_status}

# 5. 刪除租戶（連帶清除該校的使用者、學生、路線，避免外鍵孤兒資料）
@app.delete("/api/admin/school/{school_id}")
def delete_school(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SuperAdmin")),
):
    school = db.query(models.School).filter(
        models.School.school_id == school_id
    ).first()
    if not school:
        raise HTTPException(status_code=404, detail="找不到該學校記錄")

    school_name = school.school_name  # 先留存名稱，刪除後物件會失效

    # 先刪子表資料（students / routes / users），再刪學校本身
    db.query(models.Student).filter(models.Student.school_id == school_id).delete()
    db.query(models.Route).filter(models.Route.school_id == school_id).delete()
    db.query(models.User).filter(models.User.school_id == school_id).delete()
    db.delete(school)
    db.commit()
    return {"status": "success", "message": f"已徹底刪除租戶「{school_name}」及其所有關聯資料"}

# ---------------------------------------------------------
# 【现有校车线路】SchoolAdmin 线路 / 站点管理 API
# 安全原则：一切按 token 内的 school_id 过滤；操作单条记录前先校验归属本校，杜绝越权。
# ---------------------------------------------------------

class BusLineCreate(BaseModel):
    line_name: str

class BusStopCreate(BaseModel):
    stop_name: str
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    arrival_morning: Optional[str] = None
    arrival_afternoon: Optional[str] = None

class BusLineUpdate(BaseModel):
    line_name: Optional[str] = None
    arrival_school_morning: Optional[str] = None
    departure_school_afternoon: Optional[str] = None

def _serialize_stop(s: models.BusStop) -> dict:
    return {
        "stop_id": s.stop_id,
        "stop_name": s.stop_name,
        "address": s.address,
        "latitude": s.latitude,
        "longitude": s.longitude,
        "arrival_morning": s.arrival_morning,
        "arrival_afternoon": s.arrival_afternoon,
        "sequence": s.sequence,
    }

# 本校基本信息（供地图定位用）
@app.get("/api/school/me")
def get_my_school(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    school = db.query(models.School).filter(
        models.School.school_id == current_user.school_id
    ).first()
    if not school:
        raise HTTPException(status_code=404, detail="找不到您所属的学校记录")
    return {
        "school_id": school.school_id,
        "school_name": school.school_name,
        "address": school.address,
        "latitude": school.latitude,
        "longitude": school.longitude,
    }

# 获取本校所有线路（含嵌套站点）
@app.get("/api/school/lines")
def list_lines(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    lines = db.query(models.BusLine).filter(
        models.BusLine.school_id == current_user.school_id
    ).order_by(models.BusLine.line_id).all()

    return [
        {
            "line_id": ln.line_id,
            "line_name": ln.line_name,
            "arrival_school_morning": ln.arrival_school_morning,
            "departure_school_afternoon": ln.departure_school_afternoon,
            "stops": [_serialize_stop(s) for s in ln.stops],
        }
        for ln in lines
    ]

# 新建线路
@app.post("/api/school/lines")
def create_line(
    payload: BusLineCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    name = payload.line_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="线路名称不能为空")

    line = models.BusLine(line_name=name, school_id=current_user.school_id)
    db.add(line)
    db.commit()
    db.refresh(line)
    return {"status": "success", "line_id": line.line_id, "line_name": line.line_name}

# 更新线路信息（名称、到校时间、发车时间）；先校验归属本校
@app.put("/api/school/lines/{line_id}")
def update_line(
    line_id: int,
    payload: BusLineUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    line = db.query(models.BusLine).filter(
        models.BusLine.line_id == line_id,
        models.BusLine.school_id == current_user.school_id,
    ).first()
    if not line:
        raise HTTPException(status_code=404, detail="找不到该线路，或其不属于您的学校")

    if payload.line_name is not None:
        line.line_name = payload.line_name.strip() or line.line_name
    if payload.arrival_school_morning is not None:
        line.arrival_school_morning = payload.arrival_school_morning or None
    if payload.departure_school_afternoon is not None:
        line.departure_school_afternoon = payload.departure_school_afternoon or None

    db.commit()
    return {
        "status": "success",
        "line_id": line.line_id,
        "arrival_school_morning": line.arrival_school_morning,
        "departure_school_afternoon": line.departure_school_afternoon,
    }

# 删除线路（级联删除其站点）；先校验归属本校
@app.delete("/api/school/lines/{line_id}")
def delete_line(
    line_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    line = db.query(models.BusLine).filter(
        models.BusLine.line_id == line_id,
        models.BusLine.school_id == current_user.school_id,
    ).first()
    if not line:
        raise HTTPException(status_code=404, detail="找不到该线路，或其不属于您的学校")

    db.delete(line)  # cascade 连带删除其站点
    db.commit()
    return {"status": "success", "message": "线路及其站点已删除"}

# 在指定线路下新建站点；先校验该线路属本校
@app.post("/api/school/lines/{line_id}/stops")
def create_stop(
    line_id: int,
    payload: BusStopCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    line = db.query(models.BusLine).filter(
        models.BusLine.line_id == line_id,
        models.BusLine.school_id == current_user.school_id,
    ).first()
    if not line:
        raise HTTPException(status_code=404, detail="找不到该线路，或其不属于您的学校")

    name = payload.stop_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="站点名称不能为空")

    # sequence 自动取该线路末位 +1
    max_seq = db.query(models.BusStop).filter(
        models.BusStop.line_id == line_id
    ).count()

    stop = models.BusStop(
        line_id=line_id,
        stop_name=name,
        address=payload.address,
        latitude=payload.latitude,
        longitude=payload.longitude,
        arrival_morning=payload.arrival_morning,
        arrival_afternoon=payload.arrival_afternoon,
        sequence=max_seq + 1,
    )
    db.add(stop)
    db.commit()
    db.refresh(stop)
    return {"status": "success", "stop": _serialize_stop(stop)}

# 删除站点；先校验其线路属本校（join 回 BusLine 校验 school_id）
@app.delete("/api/school/stops/{stop_id}")
def delete_stop(
    stop_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    stop = db.query(models.BusStop).join(
        models.BusLine, models.BusStop.line_id == models.BusLine.line_id
    ).filter(
        models.BusStop.stop_id == stop_id,
        models.BusLine.school_id == current_user.school_id,
    ).first()
    if not stop:
        raise HTTPException(status_code=404, detail="找不到该站点，或其不属于您的学校")

    db.delete(stop)
    db.commit()
    return {"status": "success", "message": "站点已删除"}

# ---------------------------------------------------------
# 【个人信息】SchoolAdmin 头像 / 邮箱 / 密码管理
# ---------------------------------------------------------

class ProfileUpdate(BaseModel):
    email: Optional[str] = None
    avatar: Optional[str] = None          # base64 data URL
    old_password: Optional[str] = None    # 修改密码时必填
    new_password: Optional[str] = None    # 修改密码时必填

@app.get("/api/school/profile")
def get_profile(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    return {
        "username": current_user.username,
        "email": current_user.email,
        "avatar": current_user.avatar,
    }

@app.put("/api/school/profile")
def update_profile(
    payload: ProfileUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    # 更新头像
    if payload.avatar is not None:
        current_user.avatar = payload.avatar

    # 更新邮箱（检查唯一性）
    if payload.email is not None:
        email = payload.email.strip()
        if email and email != current_user.email:
            existing = db.query(models.User).filter(
                models.User.email == email,
                models.User.user_id != current_user.user_id,
            ).first()
            if existing:
                raise HTTPException(status_code=400, detail="该邮箱已被其他用户使用")
            current_user.email = email

    # 更新密码（需验证旧密码）
    if payload.new_password:
        if not payload.old_password:
            raise HTTPException(status_code=400, detail="请输入旧密码")
        if not verify_password(payload.old_password, current_user.password_hash):
            raise HTTPException(status_code=400, detail="旧密码不正确")
        current_user.password_hash = get_password_hash(payload.new_password)

    db.commit()
    return {"status": "success", "message": "个人信息已更新"}

# ---------------------------------------------------------
# 【站点编辑】更新站点信息
# ---------------------------------------------------------

class BusStopUpdate(BaseModel):
    stop_name: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    arrival_morning: Optional[str] = None
    arrival_afternoon: Optional[str] = None

@app.put("/api/school/stops/{stop_id}")
def update_stop(
    stop_id: int,
    payload: BusStopUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
):
    stop = db.query(models.BusStop).join(
        models.BusLine, models.BusStop.line_id == models.BusLine.line_id
    ).filter(
        models.BusStop.stop_id == stop_id,
        models.BusLine.school_id == current_user.school_id,
    ).first()
    if not stop:
        raise HTTPException(status_code=404, detail="找不到该站点，或其不属于您的学校")

    if payload.stop_name is not None:
        name = payload.stop_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="站点名称不能为空")
        stop.stop_name = name
    if payload.address is not None:
        stop.address = payload.address or None
    if payload.latitude is not None:
        stop.latitude = payload.latitude
    if payload.longitude is not None:
        stop.longitude = payload.longitude
    if payload.arrival_morning is not None:
        stop.arrival_morning = payload.arrival_morning or None
    if payload.arrival_afternoon is not None:
        stop.arrival_afternoon = payload.arrival_afternoon or None

    db.commit()
    return {"status": "success", "stop": _serialize_stop(stop)}

# ============================================================================
#  优化校车线路 API（第二个 tab）
# ============================================================================
from pydantic import BaseModel
from typing import Optional, List
import json as json_module

class ClusterRequest(BaseModel):
    bus_count: int
    bus_capacity: int

class OptimizeStopsRequest(BaseModel):
    max_stops_per_route: int

class StopPositionUpdate(BaseModel):
    stop_id: int
    latitude: float
    longitude: float

class StudentAddressUpdate(BaseModel):
    address: str
    latitude: float
    longitude: float


# --- 上传 Excel 学生数据 ---
@app.post("/api/school/students/upload")
async def upload_students(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")

    import openpyxl
    from io import BytesIO

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    # 清空该校已有学生数据（重新导入）
    db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id
    ).delete()
    # 同时清空优化站点和路线
    db.query(models.OptimizedStop).filter(
        models.OptimizedStop.school_id == current_user.school_id
    ).delete()
    db.query(models.OptimizedRoute).filter(
        models.OptimizedRoute.school_id == current_user.school_id
    ).delete()
    db.flush()

    # 读取表头，动态定位 name/grade/address 所在列
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    col_name = col_grade = col_address = None
    if header_row:
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            val = str(cell).strip().lower()
            if col_name is None and any(k in val for k in ("name", "student", "姓名")):
                col_name = idx
            elif col_grade is None and any(k in val for k in ("grade", "年级")):
                col_grade = idx
            elif col_address is None and any(k in val for k in ("address", "地址", "home")):
                col_address = idx
    # 找不到表头则回退到默认列顺序（name=0, grade=1, address=2）
    if col_name is None:
        col_name, col_grade, col_address = 0, 1, 2

    students_added = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col_name]:
            continue
        name = str(row[col_name]).strip()
        grade = str(row[col_grade]).strip() if col_grade is not None and row[col_grade] else ""
        address = str(row[col_address]).strip() if col_address is not None and len(row) > col_address and row[col_address] else ""

        if not name:
            continue

        student = models.Student(
            name=name,
            grade=grade,
            address=address,
            school_id=current_user.school_id,
        )
        db.add(student)
        students_added.append({"name": name, "grade": grade, "address": address})

    db.commit()
    return {"status": "success", "count": len(students_added), "students": students_added}


# --- 获取学生列表 ---
@app.get("/api/school/students")
def get_students(
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    students = db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id
    ).all()

    return [{
        "student_id": s.student_id,
        "name": s.name,
        "grade": s.grade,
        "address": s.address,
        "latitude": s.latitude,
        "longitude": s.longitude,
        "zone": s.zone,
        "assigned_stop_id": s.assigned_stop_id,
    } for s in students]


# --- 批量 Geocoding ---
@app.post("/api/school/students/geocode")
def geocode_students(
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    from optimization import geocode_address

    # 读取 Google API Key
    api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if not api_key:
        config_path = os.path.join(FRONTEND_DIR, "config.js")
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                for line in f:
                    if "GOOGLE_MAPS_API_KEY" in line and "=" in line:
                        api_key = line.split("=", 1)[1].strip().strip('"').strip("'").rstrip(";")
                        break

    students = db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id,
        models.Student.latitude.is_(None),
    ).all()

    if not students:
        # 所有学生都已有坐标
        all_students = db.query(models.Student).filter(
            models.Student.school_id == current_user.school_id
        ).all()
        return {"status": "success", "geocoded": 0, "total": len(all_students), "failed": []}

    success_count = 0
    failed = []

    for s in students:
        if not s.address:
            failed.append({"student_id": s.student_id, "name": s.name, "reason": "地址为空"})
            continue

        result = geocode_address(s.address, api_key)
        if result:
            s.latitude = result["lat"]
            s.longitude = result["lng"]
            success_count += 1
        else:
            failed.append({"student_id": s.student_id, "name": s.name, "reason": "无法解析地址"})

    db.commit()

    total = db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id
    ).count()

    return {"status": "success", "geocoded": success_count, "total": total, "failed": failed}


# --- K-Means 分区 ---
@app.post("/api/school/students/cluster")
def cluster_students(
    payload: ClusterRequest,
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    from optimization import balanced_kmeans

    students = db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id,
        models.Student.latitude.isnot(None),
    ).all()

    if not students:
        raise HTTPException(status_code=400, detail="没有已解析坐标的学生数据")

    n = len(students)
    if payload.bus_count * payload.bus_capacity < n:
        raise HTTPException(
            status_code=400,
            detail=f"运力不足：{payload.bus_count} 辆 × {payload.bus_capacity} 座 = "
                   f"{payload.bus_count * payload.bus_capacity} < {n} 名学生"
        )

    coords = [(s.latitude, s.longitude) for s in students]
    labels, centroids = balanced_kmeans(coords, payload.bus_count, payload.bus_capacity)

    # 写入分区标签
    zone_stats = {}
    for s, label in zip(students, labels):
        zone_name = f"Zone_{int(label) + 1}"
        s.zone = zone_name
        zone_stats[zone_name] = zone_stats.get(zone_name, 0) + 1

    # 清空之前的优化站点
    db.query(models.OptimizedStop).filter(
        models.OptimizedStop.school_id == current_user.school_id
    ).delete()

    db.commit()

    return {
        "status": "success",
        "zones": zone_stats,
        "students": [{
            "student_id": s.student_id,
            "name": s.name,
            "grade": s.grade,
            "latitude": s.latitude,
            "longitude": s.longitude,
            "zone": s.zone,
        } for s in students]
    }


# --- 计算最优站点 ---
@app.post("/api/school/students/optimize-stops")
def optimize_stops(
    payload: OptimizeStopsRequest,
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    from optimization import find_optimal_stops

    students = db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id,
        models.Student.zone.isnot(None),
    ).all()

    if not students:
        raise HTTPException(status_code=400, detail="请先执行分区操作")

    # 清空旧站点
    db.query(models.OptimizedStop).filter(
        models.OptimizedStop.school_id == current_user.school_id
    ).delete()
    db.flush()

    # 按区处理
    zones = sorted(set(s.zone for s in students))
    all_stops = []

    for zone in zones:
        zone_students = [s for s in students if s.zone == zone]
        zone_coords = [(s.latitude, s.longitude) for s in zone_students]

        stops, assignments = find_optimal_stops(zone_coords, payload.max_stops_per_route)

        # 存入数据库
        for seq, stop_data in enumerate(stops, start=1):
            opt_stop = models.OptimizedStop(
                school_id=current_user.school_id,
                zone=zone,
                stop_name=f"{zone}_Stop_{seq}",
                latitude=stop_data["lat"],
                longitude=stop_data["lng"],
                student_count=stop_data["student_count"],
                sequence=seq,
            )
            db.add(opt_stop)
            db.flush()

            # 更新学生分配
            for i, assign_idx in enumerate(assignments):
                if assign_idx == seq - 1:
                    zone_students[i].assigned_stop_id = opt_stop.id

            all_stops.append({
                "id": opt_stop.id,
                "zone": zone,
                "stop_name": opt_stop.stop_name,
                "latitude": stop_data["lat"],
                "longitude": stop_data["lng"],
                "student_count": stop_data["student_count"],
                "sequence": seq,
            })

    db.commit()

    return {"status": "success", "stops": all_stops}


# --- 更新站点位置（拖拽后） ---
@app.put("/api/school/students/update-stop-position")
def update_stop_position(
    payload: StopPositionUpdate,
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    stop = db.query(models.OptimizedStop).filter(
        models.OptimizedStop.id == payload.stop_id,
        models.OptimizedStop.school_id == current_user.school_id,
    ).first()

    if not stop:
        raise HTTPException(status_code=404, detail="站点不存在")

    stop.latitude = payload.latitude
    stop.longitude = payload.longitude
    db.commit()

    return {"status": "success", "stop_id": stop.id, "latitude": stop.latitude, "longitude": stop.longitude}


# --- 生成优化路线 ---
@app.post("/api/school/students/generate-routes")
def generate_optimized_routes(
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    from optimization import fetch_distance_matrix, solve_open_route
    import datetime

    def _next_monday_ts(hour, minute):
        # 用美东夏令时 (EDT = UTC-4) 计算下周一指定时刻的 UTC 时间戳
        # Google Distance Matrix API 按坐标所在地时区解读 departure_time
        edt = datetime.timezone(datetime.timedelta(hours=-4))
        now = datetime.datetime.now(edt)
        days_ahead = (7 - now.weekday()) % 7 or 7
        target = (now + datetime.timedelta(days=days_ahead)).replace(
            hour=hour, minute=minute, second=0, microsecond=0
        )
        return int(target.timestamp())

    school = db.query(models.School).filter(
        models.School.school_id == current_user.school_id
    ).first()

    if not school or not school.latitude:
        raise HTTPException(status_code=400, detail="学校坐标未设置")

    stops = db.query(models.OptimizedStop).filter(
        models.OptimizedStop.school_id == current_user.school_id
    ).all()

    if not stops:
        raise HTTPException(status_code=400, detail="请先计算优化站点")

    # 读取 API Key
    api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if not api_key:
        config_path = os.path.join(FRONTEND_DIR, "config.js")
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                for line in f:
                    if "GOOGLE_MAPS_API_KEY" in line and "=" in line:
                        api_key = line.split("=", 1)[1].strip().strip('"').strip("'").rstrip(";")
                        break

    school_point = {"lat": school.latitude, "lng": school.longitude}
    zones = sorted(set(s.zone for s in stops))

    morning_ts   = _next_monday_ts(7, 0)
    afternoon_ts = _next_monday_ts(16, 0)

    route_data    = {"morning": {}, "afternoon": {}}
    metrics_data  = {"morning": {}, "afternoon": {}}
    grand_sec_m   = 0
    grand_met_m   = 0
    grand_sec_a   = 0
    grand_met_a   = 0

    for zone in zones:
        zone_stops = [s for s in stops if s.zone == zone]
        zone_stops.sort(key=lambda s: s.sequence)

        all_points = [school_point] + [{"lat": s.latitude, "lng": s.longitude} for s in zone_stops]

        # --- 早上 ---
        time_mat_m, dist_mat_m = fetch_distance_matrix(all_points, api_key, morning_ts)
        best_perm_m, min_sec_m = solve_open_route(time_mat_m)

        # --- 下午 ---
        time_mat_a, dist_mat_a = fetch_distance_matrix(all_points, api_key, afternoon_ts)
        best_perm_a, min_sec_a = solve_open_route(time_mat_a)

        def build_ordered_stops(perm):
            idx_seq = list(perm) + [0] if perm else [0]
            ordered = []
            for idx in idx_seq:
                if idx == 0:
                    ordered.append({"name": "School", "lat": school.latitude, "lng": school.longitude})
                else:
                    s = zone_stops[idx - 1]
                    ordered.append({
                        "id": s.id,
                        "stop_name": s.stop_name,
                        "lat": s.latitude,
                        "lng": s.longitude,
                        "student_count": s.student_count,
                    })
            return ordered, idx_seq

        def build_metrics(perm, time_mat, dist_mat, min_sec):
            _, idx_seq = build_ordered_stops(perm)
            total_meters = 0
            segments = []
            for a, b in zip(idx_seq[:-1], idx_seq[1:]):
                seg_m = dist_mat[a][b]
                seg_s = time_mat[a][b]
                total_meters += seg_m
                segments.append({"km": round(seg_m / 1000, 2), "minutes": round(seg_s / 60, 1)})
            return {
                "total_km": round(total_meters / 1000, 2),
                "total_minutes": round(min_sec / 60, 1),
                "num_stops": len(zone_stops),
                "segments": segments,
            }, total_meters

        ordered_m, _ = build_ordered_stops(best_perm_m)
        ordered_a, _ = build_ordered_stops(best_perm_a)
        metrics_m, total_met_m = build_metrics(best_perm_m, time_mat_m, dist_mat_m, min_sec_m)
        metrics_a, total_met_a = build_metrics(best_perm_a, time_mat_a, dist_mat_a, min_sec_a)

        grand_sec_m += min_sec_m
        grand_met_m += total_met_m
        grand_sec_a += min_sec_a
        grand_met_a += total_met_a

        route_data["morning"][zone]   = ordered_m
        route_data["afternoon"][zone] = ordered_a
        metrics_data["morning"][zone]   = metrics_m
        metrics_data["afternoon"][zone] = metrics_a

        # 更新站点顺序（以早上路线为准）
        for new_seq, idx in enumerate(best_perm_m or [], start=1):
            zone_stops[idx - 1].sequence = new_seq

    metrics_data["morning"]["grand_total"] = {
        "total_km": round(grand_met_m / 1000, 2),
        "total_minutes": round(grand_sec_m / 60, 1),
    }
    metrics_data["afternoon"]["grand_total"] = {
        "total_km": round(grand_met_a / 1000, 2),
        "total_minutes": round(grand_sec_a / 60, 1),
    }

    # 存入数据库
    db.query(models.OptimizedRoute).filter(
        models.OptimizedRoute.school_id == current_user.school_id
    ).delete()

    opt_route = models.OptimizedRoute(
        school_id=current_user.school_id,
        route_data=json_module.dumps(route_data, ensure_ascii=False),
        metrics_data=json_module.dumps(metrics_data, ensure_ascii=False),
    )
    db.add(opt_route)
    db.commit()

    return {"status": "success", "routes": route_data, "metrics": metrics_data}


# --- 手动更新单个学生地址及坐标 ---
@app.put("/api/school/students/{student_id}/address")
def update_student_address(
    student_id: int,
    payload: StudentAddressUpdate,
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    student = db.query(models.Student).filter(
        models.Student.student_id == student_id,
        models.Student.school_id == current_user.school_id,
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    student.address = payload.address
    student.latitude = payload.latitude
    student.longitude = payload.longitude
    db.commit()
    return {"status": "success", "student_id": student_id}


# --- 清空学生数据 ---
@app.delete("/api/school/students/clear")
def clear_students(
    current_user: models.User = Depends(require_role("SchoolAdmin")),
    db: Session = Depends(get_db),
):
    db.query(models.OptimizedRoute).filter(
        models.OptimizedRoute.school_id == current_user.school_id
    ).delete()
    db.query(models.OptimizedStop).filter(
        models.OptimizedStop.school_id == current_user.school_id
    ).delete()
    db.query(models.Student).filter(
        models.Student.school_id == current_user.school_id
    ).delete()
    db.commit()
    return {"status": "success", "message": "已清空所有学生和优化数据"}


# --- 静态资源挂载（放在所有 API 路由之后，避免覆盖 /api 等路由）---
# /data：前端 fetch 的 JSON 数据文件；/frontend：HTML/JS 等前端资源
app.mount("/data", StaticFiles(directory=DATA_DIR), name="data")
app.mount("/frontend", StaticFiles(directory=FRONTEND_DIR), name="frontend")
